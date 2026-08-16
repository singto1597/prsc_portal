"""
Service ชั้นธุรกิจสำหรับระบบ Import นักเรียนจาก Excel แบบ Queue (ARQ Worker)
=============================================================================
แทนที่ `import_students_from_excel` แบบซิงโครนัส (รันยาวๆ ค้าง request) ด้วย
สถาปัตยกรรมคิว:

  POST /upload-student-excel   → เช็คคอลัมน์แบบเป๊ะ + บันทึกไฟล์ + สร้าง job (PENDING)
  POST /start-import-job/{id}  → ยิง job_id เข้า Redis (ARQ) → worker เริ่มทำงาน
  GET  /import-jobs            → ดูสถานะ / ความคืบหน้า (progress bar)

Status flow: PENDING → QUEUED → PROCESSING → COMPLETED / FAILED
"""
import json
import logging
import os
import uuid
import io
from typing import Optional, Tuple, List

import asyncpg
import openpyxl

from core.config import settings
from core.exceptions import (
    NotFoundError, ForbiddenError, ValidationError, ConflictError, ServiceUnavailableError,
)
from core.logger import AuditLogger
from core.rbac import get_role_permissions, get_role_is_admin
from models.import_schemas import (
    IMPORT_STATUS_PENDING,
    IMPORT_STATUS_QUEUED,
    IMPORT_STATUS_PROCESSING,
    IMPORT_STATUS_COMPLETED,
    IMPORT_STATUS_FAILED,
    RESTARTABLE_STATUS,
    CLAIMABLE_STATUS,
)
from services import auth_service
from services.student_service import map_role_label

logger = logging.getLogger("IMPORT_SERVICE")

# ============================================================
# 📋 คอลัมน์ของไฟล์ Excel (Format — ตรวจแบบเป๊ะ)
# ============================================================
KNOWN_COLUMNS = [
    "รหัสนักเรียน", "ห้องเรียน", "เลขที่",
    "คำนำหน้า", "ชื่อ", "นามสกุล", "ชื่อเล่น", "ตำแหน่งในห้องเรียน",
]
KNOWN_COLUMNS_SET = set(KNOWN_COLUMNS)
REQUIRED_COLUMNS = {"รหัสนักเรียน", "ห้องเรียน", "เลขที่"}

# จำนวน error รายแถวที่เก็บใน DB (กัน JSONB โตไม่มีขอบเขต)
MAX_ERROR_LOGS = 500

# 🛡️ Role ระดับโรงเรียน (ทำงานทั้งโรงเรียน ไม่ผูกห้อง/ระดับ) — สร้างได้เฉพาะผู้ดูแลทั้งโรงเรียน
# (allowed_level is None) เท่านั้น ครูระดับชั้น (allowed_level set) นำเข้า role เหล่านี้ไม่ได้
SCHOOL_WIDE_ROLES = {"admin", "teacher_council", "council_president", "council_member"}


# ============================================================
# 🧾 การอ่าน / ตรวจสอบไฟล์ Excel (Column validation แบบเป๊ะ)
# ============================================================
def validate_columns(header: list) -> dict:
    """
    ตรวจสอบหัวคอลัมน์แบบเคร่งครัด (ต้องเป๊ะ):
      1. ต้องมีคอลัมน์ที่จำเป็นครบ: รหัสนักเรียน, ห้องเรียน, เลขที่
      2. ห้ามมีคอลัมน์ซ้ำ
      3. ห้ามมีคอลัมน์ที่ไม่อยู่ใน KNOWN_COLUMNS (คอลัมน์เกิน/แปลกปลอม)

    คืน: {ชื่อคอลัมน์: index} (ข้ามคอลัมน์ว่าง)
    Raise: ValidationError พร้อมข้อความไทยที่ชี้ปัญหา
    """
    normalized = [str(c).strip() if c is not None else "" for c in header]

    if not normalized or all(c == "" for c in normalized):
        raise ValidationError("ไฟล์ว่างเปล่า — ไม่พบหัวคอลัมน์")

    seen: dict = {}
    for name in normalized:
        if name == "":
            continue
        if name in seen:
            raise ValidationError(f"พบคอลัมน์ซ้ำในไฟล์: '{name}'")
        seen[name] = True

    missing = [c for c in sorted(REQUIRED_COLUMNS) if c not in seen]
    if missing:
        raise ValidationError(
            f"ไม่พบคอลัมน์: {', '.join(missing)} — ไฟล์ต้องมีคอลัมน์ {', '.join(sorted(REQUIRED_COLUMNS))}"
        )

    unknown = [c for c in normalized if c and c not in KNOWN_COLUMNS_SET]
    if unknown:
        raise ValidationError(
            f"พบคอลัมน์ที่ไม่อยู่ในรูปแบบ: {', '.join(sorted(set(unknown)))} — รองรับเฉพาะ {', '.join(KNOWN_COLUMNS)}"
        )

    return {name: idx for idx, name in enumerate(normalized) if name}


def load_workbook_rows(file_bytes: bytes) -> Tuple[List, dict]:
    """
    อ่านไฟล์ .xlsx → (data_rows, col_index)
    - ตัดแถวว่างทิ้ง (คำนวณ total_rows = len(data_rows))
    - ตรวจคอลัมน์แบบเป๊ะ (validate_columns)
    Raise: ValidationError ถ้าอ่านไม่ได้ / คอลัมน์ผิด
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValidationError(f"อ่านไฟล์ Excel ไม่สำเร็จ — ตรวจสอบว่าเป็นไฟล์ .xlsx ที่ถูกต้อง ({e})")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValidationError("ไฟล์ว่างเปล่า")

    col_index = validate_columns(list(rows[0]))

    data_rows = []
    for r in rows[1:]:
        # ข้ามแถวที่ว่างทั้งแถว (ไม่มีข้อมูลอะไรเลย)
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        data_rows.append(r)

    return data_rows, col_index


# ============================================================
# 💾 Storage — บันทึกไฟล์ลง Server/Storage
# ============================================================
def ensure_storage_dir() -> str:
    """สร้างโฟลเดอร์เก็บไฟล์ถ้ายังไม่มี (เรียกทุกครั้งที่บันทึก — กัน deploy ลบ folder)"""
    d = settings.IMPORT_STORAGE_DIR
    os.makedirs(d, exist_ok=True)
    return d


def save_upload_file(content: bytes, original_filename: str) -> str:
    """
    บันทึกไฟล์ลง storage ด้วยชื่อสุ่ม (uuid) — กัน path traversal / ชื่อชนกัน
    คืน: absolute file_path (เก็บใน DB ใช้ worker อ่านต่อ)
    """
    ensure_storage_dir()
    stored_name = f"{uuid.uuid4().hex}.xlsx"
    file_path = os.path.join(settings.IMPORT_STORAGE_DIR, stored_name)
    with open(file_path, "wb") as f:
        f.write(content)
    return file_path


# ============================================================
# 🔁 ARQ Redis — Enqueue Job เข้าคิว
# ============================================================
async def create_arq_redis():
    """สร้าง ArqRedis (pool) จาก settings.REDIS_URL"""
    from arq.connections import RedisSettings, create_pool as arq_create_pool

    return await arq_create_pool(RedisSettings.from_dsn(settings.REDIS_URL))


async def enqueue_import_job(job_id: int) -> None:
    """ยิง job_id เข้า Redis queue เพื่อให้ ARQ Worker รับไปทำงาน"""
    redis = await create_arq_redis()
    try:
        # ชื่อฟังก์ชันต้องตรงกับ __qualname__ ของ task ใน WorkerSettings (process_student_import)
        await redis.enqueue_job("process_student_import", job_id)
        logger.info(f"✅ Enqueued import job {job_id} → Redis")
    finally:
        await redis.aclose()


# ============================================================
# 📝 สร้าง Job (Upload)
# ============================================================
async def create_import_job(
    pool: asyncpg.Pool,
    *,
    content: bytes,
    original_filename: str,
    default_password: str = "1234",
    allowed_level: Optional[str] = None,
    actor_user_id: Optional[int] = None,
    client_source: str = "web",
) -> dict:
    """
    ตรวจคอลัมน์แบบเป๊ะ → บันทึกไฟล์ลง storage → สร้าง record (status=PENDING) + audit
    (ตรวจคอลัมน์ก่อนบันทึกไฟล์ กันไฟล์ขยะค้างใน storage เมื่อคอลัมน์ผิด)
    """
    data_rows, _ = load_workbook_rows(content)  # Raise ValidationError ถ้าคอลัมน์/ไฟล์ผิด

    # ไฟล์ที่ไม่มีข้อมูลเลย (มีแค่หัวตาราง / แถวว่างล้วน) — กันส่งงานที่เปล่าประโยชน์เข้าคิว
    if not data_rows:
        raise ValidationError("ไฟล์นี้มีแค่หัวตาราง ไม่มีข้อมูลนักเรียนให้นำเข้า — กรุณาใส่ข้อมูลก่อนอัปโหลด")

    # จำกัดจำนวนแถวต่อไฟล์ (กันไฟล์ยักษ์ลาก worker หลายชั่วโมง — DoS)
    if len(data_rows) > settings.IMPORT_MAX_ROWS:
        raise ValidationError(
            f"ไฟล์มีข้อมูลเกิน {settings.IMPORT_MAX_ROWS} แถว — กรุณาแบ่งไฟล์ออกเป็นหลายไฟล์"
        )

    file_path = save_upload_file(content, original_filename)

    async with pool.acquire() as conn:
        async with conn.transaction():
            job_id = await conn.fetchval(
                """
                INSERT INTO student_import_jobs
                    (file_name, file_path, status, total_rows, default_password, allowed_level, created_by)
                VALUES ($1, $2, 'PENDING', $3, $4, $5, $6)
                RETURNING id
                """,
                original_filename, file_path, len(data_rows),
                default_password or "1234", allowed_level, actor_user_id,
            )

            # 🛡️ Audit log (กฎ: ทุก create ต้องบันทึกใน transaction เดียวกัน)
            await AuditLogger("import_service").log(
                conn=conn,
                action="UPLOAD_IMPORT_EXCEL",
                actor_identifier=str(actor_user_id) if actor_user_id else "system",
                client_source=client_source,
                user_id=actor_user_id,
                entity_type="student_import_job",
                entity_id=str(job_id),
                new_values={"file_name": original_filename, "total_rows": len(data_rows)},
            )

            row = await conn.fetchrow("SELECT * FROM student_import_jobs WHERE id = $1", job_id)

    return dict(row)


# ============================================================
# ▶️ เริ่ม Job (Start → Enqueue)
# ============================================================
async def start_import_job(
    pool: asyncpg.Pool,
    job_id: int,
    *,
    actor_user_id: Optional[int] = None,
    client_source: str = "web",
    access_scope: str = "all",
    access_level: Optional[str] = None,
) -> dict:
    """
    สั่งรันงาน: ล็อกแถว (FOR UPDATE) → ตรวจสิทธิ์/สถานะ → ตั้ง QUEUED (ล้าง error เก่า) → ยิง Redis
    ถ้า Redis ล้มเหลว → rollback (สถานะกลับเดิม) + Raise ServiceUnavailableError (503)
    """
    async with pool.acquire() as conn:
        async with conn.transaction():
            job = await conn.fetchrow(
                "SELECT * FROM student_import_jobs WHERE id = $1 FOR UPDATE", job_id
            )
            if not job:
                raise NotFoundError("ไม่พบงาน import นี้")

            # 🛡️ scope check: ครูระดับชั้นเริ่มงานได้เฉพาะของระดับชั้นตัวเอง; ครูที่ยังไม่มีระดับชั้นเริ่มไม่ได้
            if access_scope == "none":
                raise ForbiddenError("คุณยังไม่มีระดับชั้นที่รับผิดชอบ — ไม่สามารถเริ่มงานนำเข้าได้")
            if access_scope == "level" and job["allowed_level"] != access_level:
                raise ForbiddenError("งานนี้ไม่ใช่ระดับชั้นของคุณ — ไม่สามารถเริ่มได้")

            if job["status"] not in RESTARTABLE_STATUS:
                raise ConflictError(
                    f"งานนี้เริ่มไม่ได้เพราะสถานะ '{job['status']}' — รอให้เสร็จ/รอ worker ทำงานก่อน"
                )

            # ล้าง error/ความคืบหน้าเก่า (กรณี FAILED → ลองใหม่) เพื่อให้ progress bar เริ่มนับใหม่
            await conn.execute(
                """
                UPDATE student_import_jobs
                SET status = 'QUEUED', error_message = NULL, error_logs = '[]'::jsonb,
                    processed_rows = 0, imported_count = 0, skipped_count = 0,
                    completed_at = NULL, updated_at = NOW()
                WHERE id = $1
                """,
                job_id,
            )

            # 🛡️ Audit log (ภายใน transaction เดียวกับสถานะ)
            await AuditLogger("import_service").log(
                conn=conn,
                action="START_IMPORT_JOB",
                actor_identifier=str(actor_user_id) if actor_user_id else "system",
                client_source=client_source,
                user_id=actor_user_id,
                entity_type="student_import_job",
                entity_id=str(job_id),
                old_values={"status": job["status"]},
                new_values={"status": "QUEUED"},
            )

            # 🚀 ยิงเข้า Redis — ถ้าล้ม (Redis ลง) transaction rollback → สถานะกลับ PENDING
            try:
                await enqueue_import_job(job_id)
            except Exception as e:
                logger.error(f"❌ enqueue job {job_id} ล้มเหลว: {e}")
                raise ServiceUnavailableError("ระบบคิว (Redis) ไม่พร้อมใช้งาน — กรุณาลองใหม่ภายหลัง")

            row = await conn.fetchrow("SELECT * FROM student_import_jobs WHERE id = $1", job_id)

    return dict(row)


# ============================================================
# 📋 ดูรายการ Job + Progress
# ============================================================
async def list_import_jobs(
    pool: asyncpg.Pool,
    limit: int = 50,
    *,
    access_scope: str = "all",
    access_level: Optional[str] = None,
) -> list:
    """รายการงาน import เรียงใหม่สุดก่อน (มี LIMIT) — กรองตาม access scope ของผู้เรียก"""
    safe_limit = max(1, min(int(limit), 200))
    async with pool.acquire() as conn:
        if access_scope == "level":
            # ครูระดับชั้น: เห็นเฉพาะงานของระดับชั้นตัวเอง
            rows = await conn.fetch(
                """
                SELECT * FROM student_import_jobs
                WHERE allowed_level = $2
                ORDER BY created_at DESC
                LIMIT $1
                """,
                safe_limit, access_level,
            )
        elif access_scope == "none":
            # ครูที่ยังไม่มีระดับชั้น: ไม่เห็นงาน import ใดๆ เลย (upload/start ก็ถูก block แล้ว)
            rows = []
        else:
            rows = await conn.fetch(
                """
                SELECT * FROM student_import_jobs
                ORDER BY created_at DESC
                LIMIT $1
                """,
                safe_limit,
            )
    return [dict(r) for r in rows]


# ============================================================
# ⚙️ Worker — ประมวลผลงาน import (ทยอย insert + progress)
# ============================================================
class _ImportCtx:
    """Cache ต่อ 1 job — กัน SELECT ซ้ำ + กันสร้าง user/room ซ้ำในไฟล์เดียวกัน"""

    def __init__(self) -> None:
        self.room_cache: dict = {}   # room_code -> room_id
        self.user_cache: dict = {}   # username -> user_id


def _get(r: tuple, col_index: dict, name: str):
    """ดึงค่าคอลัมน์ตามชื่อจาก tuple แถว (ปลอดภัยถ้า index เกิน)"""
    i = col_index.get(name)
    return r[i] if i is not None and i < len(r) else None


def _to_int(value) -> int:
    """แปลงค่าจาก Excel → int (รับได้ทั้ง int/float/str; ว่าง = 0)"""
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    return int(s) if s else 0


async def _process_single_row(
    conn: asyncpg.Connection,
    r: tuple,
    col_index: dict,
    ctx: _ImportCtx,
    *,
    default_password: str,
    allowed_level: Optional[str],
) -> Tuple[bool, Optional[str]]:
    """
    ประมวลผล 1 แถว (สร้าง/อัปเดต users + rooms + students)
    คืน (success, error_msg) — จับเฉพาะ data error (ValueError/TypeError/UniqueViolation);
    asyncpg error อื่นๆ ปล่อยให้ลอยขึ้นไป (batch จะ fallback / job จะ FAILED)
    """
    try:
        # ⚠️ ใช้ `or ""` — ถ้าเซลล์ว่าง (None) `str(None)` = "None" (truthy!) → จะไปสร้าง user/room ปลอม
        student_id = str(_get(r, col_index, "รหัสนักเรียน") or "").strip()
        room_code = str(_get(r, col_index, "ห้องเรียน") or "").strip()
        try:
            student_no = _to_int(_get(r, col_index, "เลขที่"))
        except (ValueError, TypeError):
            return False, f"เลขที่ต้องเป็นตัวเลข (ได้ค่า: {_get(r, col_index, 'เลขที่')!r})"
        prefix = str(_get(r, col_index, "คำนำหน้า") or "").strip() or None
        first_name = str(_get(r, col_index, "ชื่อ") or "").strip()
        last_name = str(_get(r, col_index, "นามสกุล") or "").strip()
        nickname = str(_get(r, col_index, "ชื่อเล่น") or "").strip() or None
        role_label = str(_get(r, col_index, "ตำแหน่งในห้องเรียน") or "").strip()
        class_role = map_role_label(role_label)
        role_perms = get_role_permissions(class_role)
        role_is_admin = get_role_is_admin(class_role)

        if not student_id:
            return False, "ไม่มีรหัสนักเรียน"
        if not first_name and not last_name:
            return False, "ไม่มีชื่อ-นามสกุล"

        # 🛡️ กัน privilege escalation: ครูระดับชั้น (allowed_level set) สร้าง role ระดับโรงเรียน
        # (แอดมิน/ครูสภา/สภานักเรียน/ประธานสภา) ไม่ได้ — เพราะจะได้สิทธิ์จัดการทั้งโรงเรียน
        if allowed_level is not None and class_role in SCHOOL_WIDE_ROLES:
            return False, (
                f"ตำแหน่ง '{role_label}' ({class_role}) สร้างได้เฉพาะผู้ดูแลทั้งโรงเรียน — "
                f"ครูระดับชั้น ({allowed_level}) นำเข้าไม่ได้"
            )

        # 🌟 1. Room — ครูสภา/แอดมิน (school-wide) ไม่ผูกห้องเฉพาะ; ระบุห้องได้แต่ไม่บังคับ
        school_wide = class_role in SCHOOL_WIDE_ROLES
        room_id = None
        room_level = None
        if room_code:
            if room_code in ctx.room_cache:
                room_id = ctx.room_cache[room_code]
            else:
                room_id = await conn.fetchval(
                    "SELECT id FROM rooms WHERE room_code = $1 AND deleted_at IS NULL", room_code
                )
                if not room_id:
                    level = room_code.split("/")[0] if "/" in room_code else room_code
                    room_id = await conn.fetchval(
                        """
                        INSERT INTO rooms (room_code, room_name, level, room_number)
                        VALUES ($1, $2, $3, NULL)
                        RETURNING id
                        """,
                        room_code, room_code, level,
                    )
                ctx.room_cache[room_code] = room_id
            room_level = await conn.fetchval("SELECT level FROM rooms WHERE id = $1", room_id)

        # 🛡️ scope: ครูทั่วไปนำเข้าได้เฉพาะระดับชั้นตัวเอง
        if allowed_level and not school_wide and room_level != allowed_level:
            return False, f"ระดับชั้น {room_level or '-'} ไม่ตรงกับสิทธิ์ของคุณ ({allowed_level})"

        # ครูทั่วไป → staff_level = ระดับชั้นของห้องในแถวนั้น
        staff_level = room_level if class_role == "teacher" else None

        # 🌟 2. User — ใช้/สร้างตาม username (เช็ค cache ก่อนเสมอ)
        username = student_id
        if username in ctx.user_cache:
            user = ctx.user_cache[username]
        else:
            user = await conn.fetchval(
                "SELECT id FROM users WHERE username = $1 AND deleted_at IS NULL", username
            )
            if not user:
                initial_password = student_id if default_password in ("", "1234") else default_password
                hashed = auth_service.hash_password(initial_password)
                # ⚠️ ใช้ join เฉพาะค่าที่มี — ถ้า prefix ว่าง (None) f-string จะได้ "None ชื่อ นามสกุล"
                full_name = " ".join(p for p in (prefix, first_name, last_name) if p).strip() or username
                user = await conn.fetchval(
                    """
                    INSERT INTO users (username, password_hash, full_name)
                    VALUES ($1, $2, $3)
                    RETURNING id
                    """,
                    username, hashed, full_name,
                )
            ctx.user_cache[username] = user

        # 🌟 3. Student — upsert แบบ atomic กัน 2 worker/reimport ชนกันสร้างแถวซ้ำ
        if room_id is not None:
            # แถวที่มีห้อง → ON CONFLICT (ต้องมี partial unique index (room_id, student_id) WHERE deleted_at IS NULL)
            await conn.execute(
                """
                INSERT INTO students
                    (room_id, user_id, student_id, student_no, prefix,
                     first_name, last_name, nickname, class_role, staff_level,
                     is_admin, permissions)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
                ON CONFLICT (room_id, student_id) WHERE deleted_at IS NULL
                DO UPDATE SET
                    user_id = EXCLUDED.user_id,
                    student_no = EXCLUDED.student_no,
                    prefix = EXCLUDED.prefix,
                    first_name = EXCLUDED.first_name,
                    last_name = EXCLUDED.last_name,
                    nickname = EXCLUDED.nickname,
                    class_role = EXCLUDED.class_role,
                    staff_level = EXCLUDED.staff_level,
                    is_admin = EXCLUDED.is_admin,
                    permissions = EXCLUDED.permissions,
                    updated_at = NOW()
                """,
                room_id, user, student_id, student_no, prefix, first_name, last_name,
                nickname, class_role, staff_level, role_is_admin, json.dumps(role_perms),
            )
        else:
            # school-wide (room_id NULL): NULL ไม่ชนกันใน unique index → ใช้ SELECT-แล้ว-INSERT/UPDATE ตามเดิม
            student = await conn.fetchval(
                """
                SELECT id FROM students
                WHERE room_id IS NULL AND student_id = $1 AND deleted_at IS NULL
                """,
                student_id,
            )
            if student:
                await conn.execute(
                    """
                    UPDATE students
                    SET user_id = $1, student_no = $2, prefix = $3,
                        first_name = $4, last_name = $5, nickname = $6,
                        class_role = $7, staff_level = $8, is_admin = $9,
                        permissions = $10, updated_at = NOW()
                    WHERE id = $11
                    """,
                    user, student_no, prefix, first_name, last_name, nickname,
                    class_role, staff_level, role_is_admin, json.dumps(role_perms), student,
                )
            else:
                await conn.execute(
                    """
                    INSERT INTO students
                        (room_id, user_id, student_id, student_no, prefix,
                         first_name, last_name, nickname, class_role, staff_level,
                         is_admin, permissions)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
                    """,
                    room_id, user, student_id, student_no, prefix, first_name, last_name,
                    nickname, class_role, staff_level, role_is_admin, json.dumps(role_perms),
                )

        return True, None
    except (ValueError, TypeError, asyncpg.UniqueViolationError) as e:
        return False, f"ข้อมูลผิดรูปแบบ/ซ้ำ ({e})"


async def _claim_job(pool: asyncpg.Pool, job_id: int) -> Optional[dict]:
    """
    worker ครอบงาน: ล็อกแถว → ตรวจสถานะ (PENDING/QUEUED) → ตั้ง PROCESSING + audit
    คืน job dict (พร้อม file_path) หรือ None ถ้าจบ/ถูกครอบไปแล้ว
    """
    async with pool.acquire() as conn:
        async with conn.transaction():
            job = await conn.fetchrow(
                "SELECT * FROM student_import_jobs WHERE id = $1 FOR UPDATE", job_id
            )
            if not job:
                logger.warning(f"⚠️ Job {job_id} ไม่มีอยู่ — ข้าม")
                return None
            if job["status"] not in CLAIMABLE_STATUS:
                logger.info(f"ℹ️ Job {job_id} จบ/ถูกครอบไปแล้ว (status={job['status']}) — ข้าม")
                return None

            await conn.execute(
                """
                UPDATE student_import_jobs
                SET status = 'PROCESSING', started_at = COALESCE(started_at, NOW()), updated_at = NOW()
                WHERE id = $1
                """,
                job_id,
            )
            await AuditLogger("import_service").log(
                conn=conn,
                action="PROCESS_IMPORT_JOB",
                actor_identifier=str(job["created_by"]) if job["created_by"] else "worker",
                client_source="worker",
                user_id=job["created_by"],
                entity_type="student_import_job",
                entity_id=str(job_id),
                old_values={"status": job["status"]},
                new_values={"status": "PROCESSING"},
            )
    return dict(job)


async def _set_job_result(
    pool: asyncpg.Pool,
    job_id: int,
    *,
    status: str,
    error_message: Optional[str] = None,
    total: Optional[int] = None,
    imported: Optional[int] = None,
    skipped: Optional[int] = None,
    errors: Optional[list] = None,
) -> None:
    """จบงาน: เขียนผลลัพธ์ (COMPLETED/FAILED) + error_logs + audit ภายใน transaction เดียว"""
    error_logs = (errors or [])[:MAX_ERROR_LOGS]

    async with pool.acquire() as conn:
        async with conn.transaction():
            job = await conn.fetchrow(
                "SELECT status, created_by, file_path FROM student_import_jobs WHERE id = $1", job_id
            )
            if not job:
                return
            old_status = job["status"]

            if status == IMPORT_STATUS_FAILED:
                await conn.execute(
                    """
                    UPDATE student_import_jobs
                    SET status = 'FAILED', error_message = $2, error_logs = $3::jsonb,
                        completed_at = NOW(), updated_at = NOW()
                    WHERE id = $1
                    """,
                    job_id, error_message, json.dumps(error_logs, ensure_ascii=False),
                )
            else:
                await conn.execute(
                    """
                    UPDATE student_import_jobs
                    SET status = 'COMPLETED', total_rows = $2, processed_rows = $2,
                        imported_count = $3, skipped_count = $4, error_logs = $5::jsonb,
                        error_message = NULL,   -- ล้าง error เก่า (กัน UI โชว์ค้างหลังเสร็จ)
                        completed_at = NOW(), updated_at = NOW()
                    WHERE id = $1
                    """,
                    job_id, total, imported, skipped, json.dumps(error_logs, ensure_ascii=False),
                )

    # งานเสร็จแล้ว → ลบไฟล์ออกจาก storage (ประหยัดพื้นที่; FAILED เก็บไว้ให้ลองใหม่)
    if status == IMPORT_STATUS_COMPLETED and job.get("file_path"):
        try:
            os.remove(job["file_path"])
        except OSError as e:
            logger.warning(f"⚠️ ลบไฟล์ storage ไม่สำเร็จ: {job['file_path']} ({e})")

            await AuditLogger("import_service").log(
                conn=conn,
                action="COMPLETE_IMPORT_JOB" if status == IMPORT_STATUS_COMPLETED else "FAIL_IMPORT_JOB",
                actor_identifier=str(job["created_by"]) if job["created_by"] else "worker",
                client_source="worker",
                user_id=job["created_by"],
                entity_type="student_import_job",
                entity_id=str(job_id),
                status="success" if status == IMPORT_STATUS_COMPLETED and not error_logs else
                       ("partial" if status == IMPORT_STATUS_COMPLETED else "error"),
                error_detail=error_message,
                old_values={"status": old_status},
                new_values={"status": status, **({"error_message": error_message} if error_message else {})},
            )


async def process_import_job(pool: asyncpg.Pool, job_id: int) -> dict:
    """
    ⚙️ ฟังก์ชันหลักของ Worker — ครอบงาน → อ่านไฟล์ → ทยอย insert เป็น batch → update progress → จบ

    - แต่ละ batch เป็น transaction แยก (กันล็อกยาว + commit progress เป็นระยะ)
    - อัปเดต processed_rows ทุก batch → เอาไปทำ progress bar
    - แถวที่ข้อมูลผิด → ข้าม (บันทึกลง error_logs) ; ไฟล์พัง/ข้อผิดพลาดร้ายแรง → FAILED
    """
    job = await _claim_job(pool, job_id)
    if job is None:
        return {"status": "SKIPPED", "reason": "already_finished_or_claimed"}

    file_path = job["file_path"]
    default_password = job["default_password"] or "1234"
    allowed_level = job["allowed_level"]
    created_by = job["created_by"]
    imported = 0
    skipped = 0
    errors: List[str] = []

    try:
        # 1. อ่านไฟล์จาก storage
        try:
            with open(file_path, "rb") as f:
                content = f.read()
        except FileNotFoundError:
            # แสดงแค่ basename — อย่า leak path เต็มบน server ออกไปยังผู้ใช้
            raise ValidationError(f"ไม่พบไฟล์บน storage: {os.path.basename(file_path)}")
        except OSError as e:
            raise ValidationError(f"อ่านไฟล์จาก storage ไม่สำเร็จ: {e}")

        # 2. แยกแถว + ตรวจคอลัมน์ (เป๊ะ)
        data_rows, col_index = load_workbook_rows(content)
        total = len(data_rows)

        # 3. ทยอย insert เป็น batch
        batch_size = settings.IMPORT_BATCH_SIZE
        ctx = _ImportCtx()

        async with pool.acquire() as conn:
            for start in range(0, total, batch_size):
                chunk = data_rows[start:start + batch_size]

                # 📸 snapshot ตัวนับก่อนลอง batch — ถ้า batch rollback จะได้คืนค่า (กันนับซ้ำ)
                snapshot_imported, snapshot_skipped, snapshot_errors_len = imported, skipped, len(errors)
                try:
                    # Batch ปกติ: transaction เดียวทั้ง chunk
                    async with conn.transaction():
                        for idx, r in enumerate(chunk):
                            ok, err = await _process_single_row(
                                conn, r, col_index, ctx,
                                default_password=default_password, allowed_level=allowed_level,
                            )
                            row_no = start + idx + 1
                            if ok:
                                imported += 1
                            else:
                                skipped += 1
                                errors.append(f"แถว {row_no}: {err}")
                except asyncpg.PostgresError as e:
                    # Batch พัง (rollback ทั้ง chunk) → คืน counter + เคลียร์ cache
                    # (IDs ที่ insert ใน batch ที่พังถูก rollback ไปแล้ว → cache เก่า = phantom ID)
                    imported, skipped = snapshot_imported, snapshot_skipped
                    del errors[snapshot_errors_len:]
                    ctx.room_cache.clear()
                    ctx.user_cache.clear()
                    logger.error(f"⚠️ batch job {job_id} พัง ({e}) — fallback ทีละแถว")
                    for idx, r in enumerate(chunk):
                        row_no = start + idx + 1
                        try:
                            async with conn.transaction():
                                ok, err = await _process_single_row(
                                    conn, r, col_index, ctx,
                                    default_password=default_password, allowed_level=allowed_level,
                                )
                        except Exception as row_err:
                            ok, err = False, f"ไม่สามารถบันทึกแถวนี้ได้ ({row_err})"
                        if ok:
                            imported += 1
                        else:
                            skipped += 1
                            errors.append(f"แถว {row_no}: {err}")

                # 4. อัปเดต progress ทุก batch (หลังจาก batch commit แล้ว)
                await conn.execute(
                    """
                    UPDATE student_import_jobs
                    SET processed_rows = $2, updated_at = NOW()
                    WHERE id = $1
                    """,
                    job_id, start + len(chunk),
                )

    except ValidationError as e:
        logger.error(f"❌ Job {job_id} invalid: {e}")
        await _set_job_result(pool, job_id, status=IMPORT_STATUS_FAILED, error_message=str(e), errors=errors)
        return {"status": "FAILED", "reason": str(e)}
    except Exception as e:
        logger.exception(f"❌ Job {job_id} พังแบบไม่คาดคิด: {e}")
        await _set_job_result(pool, job_id, status=IMPORT_STATUS_FAILED, error_message=str(e), errors=errors)
        return {"status": "FAILED", "reason": str(e)}
    else:
        # 5. สำเร็จ — เขียนผลรวม + COMPLETED
        await _set_job_result(
            pool, job_id,
            status=IMPORT_STATUS_COMPLETED,
            total=total, imported=imported, skipped=skipped, errors=errors,
        )
        logger.info(f"✅ Job {job_id} COMPLETED — total={total} imported={imported} skipped={skipped}")
        return {"status": "COMPLETED", "total_rows": total, "imported": imported, "skipped": skipped, "errors": errors[:50]}


# ============================================================
# 🔧 Recovery — กู้คืนงานค้างเมื่อ worker restart
# ============================================================
async def recover_stuck_jobs(pool: asyncpg.Pool) -> None:
    """
    เรียกตอน worker เริ่มต้น: งานที่ค้างนานเกิน → reset กลับ QUEUED + ยิงคิวใหม่
    - PROCESSING ค้าง > 35 นาที: worker ก่อนหน้าตายกลางคัน
    - QUEUED ค้าง > 35 นาที: Redis หายไป (job ที่ start แล้วแต่คิวหาย) — กันติดค้างถาวร
      (QUEUED ถูกกันออกจาก RESTARTABLE_STATUS แล้ว ต้องพึ่ง recovery นี้แทน)
    """
    async with pool.acquire() as conn:
        async with conn.transaction():
            rows = await conn.fetch(
                """
                SELECT id, created_by, status FROM student_import_jobs
                WHERE status IN ('PROCESSING', 'QUEUED')
                  AND updated_at < NOW() - make_interval(mins => $1)
                FOR UPDATE
                """,
                settings.IMPORT_RECOVERY_STALE_MINUTES,
            )
            for r in rows:
                await conn.execute(
                    "UPDATE student_import_jobs SET status = 'QUEUED', updated_at = NOW() WHERE id = $1",
                    r["id"],
                )
                await AuditLogger("import_service").log(
                    conn=conn,
                    action="RECOVER_IMPORT_JOB",
                    actor_identifier=str(r["created_by"]) if r["created_by"] else "worker",
                    client_source="worker",
                    user_id=r["created_by"],
                    entity_type="student_import_job",
                    entity_id=str(r["id"]),
                    old_values={"status": r["status"]},
                    new_values={"status": "QUEUED"},
                )

    for r in rows:
        try:
            await enqueue_import_job(r["id"])
        except Exception as e:
            logger.error(f"⚠️ re-enqueue job {r['id']} หลัง recovery ล้มเหลว: {e}")
