# === Integration Tests: Import นักเรียนจาก Excel แบบ Queue (ARQ Worker) ===
# ครอบคลุม: upload + ตรวจคอลัมน์เป๊ะ → start (enqueue Redis) → worker ทยอย insert + progress
# ตามกฎ testing.md: Deep DB verification ทุกกรณี + mock Redis (ไม่เชื่อมจริง) + random ID กัน state รั่ว
import io
import json
import os
import random
import re
import zipfile

import pytest
import pytest_asyncio
import asyncpg
import openpyxl
from openpyxl import Workbook
from unittest.mock import patch, AsyncMock

from services import auth_service, import_service


# === Fixtures & Setup ===

def make_xlsx_bytes(rows, header=None):
    """สร้างไฟล์ .xlsx ในหน่วยความจำ (bytes) — ตาม Format ที่ระบบรองรับ"""
    wb = Workbook()
    ws = wb.active
    ws.append(
        header
        or ["รหัสนักเรียน", "ห้องเรียน", "เลขที่", "คำนำหน้า", "ชื่อ", "นามสกุล", "ชื่อเล่น", "ตำแหน่งในห้องเรียน"]
    )
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def make_xlsx_bytes_with_float_student_id(rows, header=None):
    """สร้างไฟล์ .xlsx ที่ cell รหัสนักเรียน (คอลัมน์ A) ถูกจัดเก็บเป็น float (<v>40000.0</v>) ใน XML —
    จำลองไฟล์จากเครื่องมืออื่น (Excel/Google Sheets/macro ที่เก็บเลขรหัสแบบทศนิยม) ที่ openpyxl
    อ่านคืนเป็น float 40000.0 → ต้อง import เป็น '40000' ไม่ใช่ '40000.0'.
    แก้เฉพาะคอลัมน์ A (r=\"A<row>\") — กันทำลาย <v> ของ shared-string index ในคอลัมน์อื่น"""
    data = make_xlsx_bytes(rows=rows, header=header)
    zin = zipfile.ZipFile(io.BytesIO(data))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/"):
                content = re.sub(
                    br'(<c r="A\d+"[^>]*>)(<v>)(\d+)(</v>)',
                    br"\1\2\3.0\4",
                    content,
                )
            zout.writestr(item, content)
    return out.getvalue()


@pytest_asyncio.fixture
async def admin_user(db_pool):
    """admin (school-wide, room_id NULL) — มี MANAGE_STUDENTS ผ่าน is_admin"""
    username = f"adm{random.randint(100000, 999999)}"  # ≤10 chars (student_id VARCHAR(10))
    uid = await auth_service.register_user(
        db_pool, username, "1234", "แอดมิน เทส", username, "", 0, "admin"
    )
    return uid, username, "1234"


@pytest_asyncio.fixture
async def student_user(db_pool):
    """นักเรียนธรรมดา — ไม่มี MANAGE_STUDENTS"""
    username = f"stu{random.randint(100000, 999999)}"
    # clean_database truncates rooms ก่อนทุก test → ต้องสร้างห้องก่อน register_user
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO rooms (room_code, room_name, level, room_number)
            VALUES ('ม.4/1', 'ม.4/1', 'ม.4', 1)
            ON CONFLICT (room_code) DO NOTHING
            """
        )
    # student_id = username (9 ตัว) — ต้อง ≤10 ตัว (VARCHAR(10)) ใช้ f"ID{username}" ไม่ได้
    uid = await auth_service.register_user(
        db_pool, username, "1234", "นักเรียน เทส", username, "ม.4/1", 1, "student"
    )
    return uid, username, "1234"


def login_token(client, username, password):
    res = client.post("/api/auth/login", json={"username": username, "password": password})
    assert res.status_code == 200, res.text
    return res.json()["access_token"]


def upload_excel(client, token, file_bytes, filename="students.xlsx", default_password="1234"):
    """POST /api/upload-student-excel (sync — TestClient)"""
    return client.post(
        "/api/upload-student-excel",
        files={"file": (filename, file_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
        params={"default_password": default_password},
    )


async def upload_and_get_job(client, token, rows, **kwargs):
    res = upload_excel(client, token, make_xlsx_bytes(rows=rows), **kwargs)
    assert res.status_code == 200, res.text
    return res.json()["id"]


# === Section 1: Upload (Happy path + Deep DB) ===

@pytest.mark.asyncio
async def test_upload_creates_pending_job_and_saves_file(client, db_pool, admin_user):
    """อัปโหลดไฟล์ถูกต้อง → 200 + job status=PENDING + total_rows + ไฟล์บันทึกใน storage"""
    _, username, password = admin_user
    token = login_token(client, username, password)

    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "ชาย", ""],
        ["47002", "ม.4/1", 2, "นางสาว", "สมหญิง", "รักเรียน", "หญิง", "หัวหน้าห้อง"],
    ]
    res = upload_excel(client, token, make_xlsx_bytes(rows=rows))
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["status"] == "PENDING"
    assert data["total_rows"] == 2
    assert data["processed_rows"] == 0
    assert data["imported_count"] == 0
    assert data["error_logs"] == [], "error_logs (JSONB) ต้อง parse เป็น [] — ไม่ใช่ ['[', ']']"
    assert data["progress_percent"] == 0
    assert data["file_name"] == "students.xlsx"

    # 🔍 Deep DB verification
    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT * FROM student_import_jobs WHERE id = $1", data["id"])
        assert job is not None
        assert job["status"] == "PENDING"
        assert job["total_rows"] == 2
        assert job["processed_rows"] == 0
        # ไฟล์ถูกบันทึกจริงบน storage
        assert os.path.exists(job["file_path"]), "ไฟล์ต้องถูกบันทึกลง storage"
        assert job["created_by"] is not None

    # Audit log ถูกเขียน (กฎ: ทุก create ต้อง audit)
    async with db_pool.acquire() as conn:
        audit = await conn.fetchval(
            "SELECT count(*) FROM audit_logs WHERE action = 'UPLOAD_IMPORT_EXCEL'"
        )
        assert audit == 1


@pytest.mark.asyncio
async def test_upload_empty_row_skipped_in_total(client, admin_user):
    """แถวว่างทั้งแถวไม่ถูกนับเป็น total_rows"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "ชาย", ""],
        [None, None, None, None, None, None, None, None],  # แถวว่าง
        ["47002", "ม.4/1", 2, "นางสาว", "สมหญิง", "รักเรียน", "หญิง", ""],
    ]
    res = upload_excel(client, token, make_xlsx_bytes(rows=rows))
    assert res.status_code == 200
    assert res.json()["total_rows"] == 2


# === Section 2: Upload — Column validation (ต้องเป๊ะ) ===

@pytest.mark.asyncio
@pytest.mark.parametrize("header, expected_fragment", [
    (["ห้องเรียน", "เลขที่"], "ไม่พบคอลัมน์: รหัสนักเรียน"),
    (["รหัสนักเรียน", "เลขที่"], "ไม่พบคอลัมน์: ห้องเรียน"),
    (["รหัสนักเรียน", "ห้องเรียน"], "ไม่พบคอลัมน์: เลขที่"),
])
async def test_upload_missing_required_column(client, admin_user, header, expected_fragment):
    """ขาดคอลัมน์ที่จำเป็น → 400 + ข้อความไทยชี้ชัด"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    res = upload_excel(client, token, make_xlsx_bytes(rows=[["47001", "ม.4/1", 1]], header=header))
    assert res.status_code == 400, res.text
    assert expected_fragment in res.json()["detail"]


@pytest.mark.asyncio
async def test_upload_rejects_unknown_column(client, admin_user):
    """คอลัมน์เกินจากที่รองรับ (strict — ต้องเป๊ะ) → 400"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    header = ["รหัสนักเรียน", "ห้องเรียน", "เลขที่", "หมายเหตุ"]
    res = upload_excel(client, token, make_xlsx_bytes(rows=[["47001", "ม.4/1", 1, "x"]], header=header))
    assert res.status_code == 400, res.text
    assert "หมายเหตุ" in res.json()["detail"]
    assert "รองรับเฉพาะ" in res.json()["detail"]


@pytest.mark.asyncio
async def test_upload_rejects_duplicate_column(client, admin_user):
    """คอลัมน์ซ้ำ → 400"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    header = ["รหัสนักเรียน", "ห้องเรียน", "เลขที่", "รหัสนักเรียน"]
    res = upload_excel(client, token, make_xlsx_bytes(rows=[["47001", "ม.4/1", 1, "47001"]], header=header))
    assert res.status_code == 400, res.text
    assert "ซ้ำ" in res.json()["detail"]


@pytest.mark.asyncio
async def test_upload_rejects_non_excel(client, admin_user):
    """ไฟล์ไม่ใช่ .xlsx (นามสกุล) → 400"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    res = upload_excel(client, token, b"not an excel file", filename="file.txt")
    assert res.status_code == 400
    assert ".xlsx" in res.json()["detail"]


@pytest.mark.asyncio
async def test_upload_rejects_corrupt_xlsx(client, admin_user):
    """ไฟล์ .xlsx ปลอม/เสีย (เปิดไม่ได้) → 400"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    res = upload_excel(client, token, b"PK not really a zip", filename="fake.xlsx")
    assert res.status_code == 400, res.text
    assert "อ่านไฟล์" in res.json()["detail"]


# === Section 3: Upload — Auth & RBAC ===

@pytest.mark.asyncio
async def test_upload_requires_auth(client):
    """ไม่มี token → 401"""
    res = upload_excel(client, "", make_xlsx_bytes(rows=[["47001", "ม.4/1", 1]]))
    assert res.status_code == 401


@pytest.mark.asyncio
async def test_upload_forbidden_without_manage_students(client, student_user):
    """นักเรียนธรรมดา (ไม่มี MANAGE_STUDENTS) → 403"""
    _, username, password = student_user
    token = login_token(client, username, password)
    res = upload_excel(client, token, make_xlsx_bytes(rows=[["47001", "ม.4/1", 1]]))
    assert res.status_code == 403


# === Section 4: Start Job (enqueue to Redis — mock) ===

@pytest.mark.asyncio
async def test_start_job_enqueues_and_sets_queued(client, db_pool, admin_user):
    """start → 200 status=QUEUED + mock enqueue ถูกเรียกด้วย job_id + Deep DB"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])

    with patch("services.import_service.enqueue_import_job", new_callable=AsyncMock) as mock_enqueue:
        res = client.post(f"/api/start-import-job/{job_id}",
                          headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 200, res.text
    assert res.json()["status"] == "QUEUED"
    mock_enqueue.assert_awaited_once_with(job_id)

    # 🔍 Deep DB verification
    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT status FROM student_import_jobs WHERE id = $1", job_id)
        assert job["status"] == "QUEUED"


@pytest.mark.asyncio
async def test_start_job_twice_conflict(client, db_pool, admin_user):
    """start ครั้งที่ 2 ขณะ QUEUED → 409 (กันยิงคิวซ้ำ)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])

    with patch("services.import_service.enqueue_import_job", new_callable=AsyncMock):
        r1 = client.post(f"/api/start-import-job/{job_id}", headers={"Authorization": f"Bearer {token}"})
        assert r1.status_code == 200
        r2 = client.post(f"/api/start-import-job/{job_id}", headers={"Authorization": f"Bearer {token}"})
    assert r2.status_code == 409, r2.text
    assert "QUEUED" in r2.json()["detail"]


@pytest.mark.asyncio
async def test_start_job_not_found(client, admin_user):
    """job_id ไม่มีในระบบ → 404"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    res = client.post("/api/start-import-job/999999", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 404


@pytest.mark.asyncio
async def test_start_job_redis_down_rolls_back(client, db_pool, admin_user):
    """Redis ล้มตอน enqueue → 503 + transaction rollback (สถานะกลับ PENDING)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])

    with patch("services.import_service.enqueue_import_job",
               side_effect=RuntimeError("Redis connection refused")):
        res = client.post(f"/api/start-import-job/{job_id}", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 503, res.text

    # 🔍 Deep DB: สถานะไม่ติด QUEUED ค้าง (rollback กลับ PENDING)
    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT status FROM student_import_jobs WHERE id = $1", job_id)
        assert job["status"] == "PENDING"


# === Section 5: Worker — ประมวลผลจริง (เรียก service ตรง, ไม่ใช้ Redis) ===

@pytest.mark.asyncio
async def test_worker_completes_job_and_creates_students(client, db_pool, admin_user):
    """worker ทำงาน: สร้าง user/room/student ครบ + status COMPLETED + counts ถูกต้อง"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "ชาย", ""],
        ["47002", "ม.4/1", 2, "นางสาว", "สมหญิง", "รักเรียน", "หญิง", "หัวหน้าห้อง"],
    ]
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 2
    assert result["skipped"] == 0

    # 🔍 Deep DB: job จบ + progress เต็ม 100
    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT * FROM student_import_jobs WHERE id = $1", job_id)
        assert job["status"] == "COMPLETED"
        assert job["processed_rows"] == job["total_rows"] == 2
        assert job["imported_count"] == 2
        assert job["skipped_count"] == 0
        assert job["completed_at"] is not None

        # students ถูกสร้าง 2 คน (scope เฉพาะคนที่ import — admin_user fixture ก็สร้าง student ของตัวเองไว้)
        students = await conn.fetch(
            "SELECT * FROM students WHERE student_id IN ('47001', '47002') ORDER BY student_id"
        )
        assert len(students) == 2
        assert {s["student_id"] for s in students} == {"47001", "47002"}
        # หัวหน้าห้อง → class_role + permissions ถูกตั้งจาก roles.json
        head = next(s for s in students if s["student_id"] == "47002")
        assert head["class_role"] == "class_president"
        assert "RECEIVE_ISSUES" in json.loads(head["permissions"])

        # room ถูกสร้างอัตโนมัติ + ระดับชั้นถูก
        room = await conn.fetchrow("SELECT * FROM rooms WHERE room_code = 'ม.4/1'")
        assert room is not None
        assert room["level"] == "ม.4"

        # users ถูกสร้างด้วยรหัสเริ่มต้น = เลขรหัสนักเรียน
        user = await conn.fetchrow("SELECT * FROM users WHERE username = '47001'")
        assert user is not None
        assert auth_service.verify_password("47001", user["password_hash"]) is True


@pytest.mark.asyncio
async def test_worker_imports_float_student_id_without_dot_zero(client, db_pool, admin_user):
    """รหัสนักเรียนที่ไฟล์เก็บเป็น float (40000.0) → import เป็น '40000' ไม่ใช่ '40000.0'
    (student_id + username + รหัสผ่านเริ่มต้นต้องไม่มี .0 ต่อท้าย)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [
        [40000, "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""],
        [47001, "ม.4/1", 2, "นางสาว", "สมหญิง", "รักเรียน", "", ""],
    ]
    res = upload_excel(client, token, make_xlsx_bytes_with_float_student_id(rows=rows))
    assert res.status_code == 200, res.text
    job_id = res.json()["id"]

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 2
    assert result["skipped"] == 0
    assert result["errors"] == [], result["errors"]

    # 🔍 Deep DB: student_id ต้องเป็น '40000' ไม่ใช่ '40000.0' + user เข้าระบบด้วย 40000/40000
    async with db_pool.acquire() as conn:
        students = await conn.fetch(
            "SELECT student_id FROM students WHERE student_id IN ('40000', '47001') ORDER BY student_id"
        )
        assert [s["student_id"] for s in students] == ["40000", "47001"]
        user = await conn.fetchrow("SELECT * FROM users WHERE username = '40000'")
        assert user is not None
        assert auth_service.verify_password("40000", user["password_hash"]) is True


@pytest.mark.asyncio
async def test_worker_skips_already_claimed_job(client, db_pool, admin_user):
    """เรียก worker 2 รอบ → รอบที่ 2 SKIPPED (กันประมวลผลซ้ำ) + ไม่มี student ซ้ำ"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])

    r1 = await import_service.process_import_job(db_pool, job_id)
    assert r1["status"] == "COMPLETED"
    r2 = await import_service.process_import_job(db_pool, job_id)
    assert r2["status"] == "SKIPPED"

    async with db_pool.acquire() as conn:
        students = await conn.fetchval("SELECT count(*) FROM students WHERE student_id = '47001'")
        assert students == 1


@pytest.mark.asyncio
async def test_worker_handles_bad_rows_but_imports_good_ones(client, db_pool, admin_user):
    """แถวข้อมูลผิด (ไม่มีรหัส / เลขที่แปลก) → ข้าม + ลง error_logs; แถวดีนำเข้าปกติ"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""],
        [None, "ม.4/1", 2, "นาย", "ไม่มีรหัส", "ไม่ได้ใส่", "", ""],           # ไม่มีรหัสนักเรียน
        ["47002", "ม.4/1", "abc", "นางสาว", "เลขที่ผิด", "แบบฟอร์ม", "", ""],  # เลขที่ไม่ใช่ตัวเลข
        ["47003", "ม.4/1", 3, "นาย", "คนที่สาม", "โอเค", "", ""],
    ]
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 2
    assert result["skipped"] == 2

    # 🔍 Deep DB: error_logs มี 2 รายการ + นับถูก
    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT * FROM student_import_jobs WHERE id = $1", job_id)
        assert job["imported_count"] == 2
        assert job["skipped_count"] == 2
        error_logs = json.loads(job["error_logs"])
        assert len(error_logs) == 2
        assert any("ไม่มีรหัสนักเรียน" in e for e in error_logs)
        assert any("เลขที่" in e for e in error_logs)
        # นำเข้าเฉพาะ 2 แถวดี (admin_user fixture ก็สร้าง student ของตัวเองไว้ ไม่นับ)
        students = await conn.fetchval(
            "SELECT count(*) FROM students WHERE student_id IN ('47001', '47003')"
        )
        assert students == 2


@pytest.mark.asyncio
async def test_worker_skips_template_sample_rows(client, db_pool, admin_user):
    """🛡️ แถวตัวอย่างจาก Template (รหัสขึ้นต้น 000) ต้องถูกข้าม — อัปโหลด Template ทั้งไฟล์ ไม่สร้าง account ปลอม"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [
        ["00001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""],          # แถวตัวอย่าง 000xx
        ["47001", "ม.4/1", 1, "นาย", "คนจริง", "ใจดี", "", ""],
        ["00002", "ม.4/2", 1, "นางสาว", "สมหญิง", "รักเรียน", "", "หัวหน้าห้อง"],  # แถวตัวอย่าง 000xx
    ]
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 1
    assert result["skipped"] == 2

    # 🔍 Deep DB: ต้องไม่มี user/student ที่รหัสขึ้นต้น 000 (00001/00002) ถูกสร้าง
    async with db_pool.acquire() as conn:
        users = await conn.fetchval(
            "SELECT count(*) FROM users WHERE username IN ('00001', '00002')"
        )
        students = await conn.fetchval(
            "SELECT count(*) FROM students WHERE student_id IN ('00001', '00002')"
        )
        assert users == 0, "แถวตัวอย่าง (000xx) ต้องไม่ถูกสร้างเป็น user"
        assert students == 0, "แถวตัวอย่าง (000xx) ต้องไม่ถูกสร้างเป็น student"
        error_logs = json.loads(
            await conn.fetchval("SELECT error_logs FROM student_import_jobs WHERE id = $1", job_id)
        )
        assert any("000" in e for e in error_logs)


@pytest.mark.asyncio
async def test_worker_failed_when_file_missing(client, db_pool, admin_user):
    """ไฟล์ถูกลบจาก storage → worker จบ FAILED + error_message"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])

    async with db_pool.acquire() as conn:
        path = await conn.fetchval("SELECT file_path FROM student_import_jobs WHERE id = $1", job_id)
    os.remove(path)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "FAILED"

    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT * FROM student_import_jobs WHERE id = $1", job_id)
        assert job["status"] == "FAILED"
        assert job["error_message"] and "ไฟล์" in job["error_message"]
        assert job["completed_at"] is not None


@pytest.mark.asyncio
async def test_teacher_import_scoped_to_own_level(client, db_pool):
    """ครูทั่วไป (ระดับ ม.4) นำเข้าได้เฉพาะแถวห้อง ม.4 — แถว ม.5 ต้องถูกข้าม"""
    username = f"tch{random.randint(100000, 999999)}"
    uid = await auth_service.register_user(
        db_pool, username, "1234", "ครู เทส", username, "", 0, "teacher"
    )
    # register_user ไม่ตั้ง staff_level → ตั้งตรง (ครู ม.4 ดูแลแค่ ม.4)
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE students SET staff_level = 'ม.4' WHERE user_id = $1", uid)

    token = login_token(client, username, "1234")
    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""],
        ["48001", "ม.5/1", 1, "นาย", "อีกคน", "นอกสาย", "", ""],   # ม.5 — นอกระดับครู
    ]
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 1
    assert result["skipped"] == 1

    # 🔍 Deep DB: มีแค่นักเรียน ม.4/1
    async with db_pool.acquire() as conn:
        students = await conn.fetch(
            "SELECT s.student_id, r.level FROM students s JOIN rooms r ON r.id = s.room_id"
        )
        assert len(students) == 1
        assert students[0]["student_id"] == "47001"
        assert students[0]["level"] == "ม.4"


@pytest.mark.asyncio
async def test_reimport_updates_existing_students(client, db_pool, admin_user):
    """นำเข้าไฟล์เดิมซ้ำ → อัปเดต student เดิม (ไม่สร้างซ้ำ)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]
    job_id = await upload_and_get_job(client, token, rows)
    await import_service.process_import_job(db_pool, job_id)

    # นำเข้าอีกครั้ง (ไฟล์ใหม่)
    rows2 = [["47001", "ม.4/1", 1, "นาย", "สมชาย", "เปลี่ยนนามสกุล", "", "หัวหน้าห้อง"]]
    job_id2 = await upload_and_get_job(client, token, rows2)
    await import_service.process_import_job(db_pool, job_id2)

    async with db_pool.acquire() as conn:
        # นับเฉพาะ 47001 (admin_user fixture ก็สร้าง user/student ของตัวเอง — ไม่นับ)
        users = await conn.fetchval("SELECT count(*) FROM users WHERE username = '47001'")
        students = await conn.fetchval("SELECT count(*) FROM students WHERE student_id = '47001'")
        assert users == 1, "user ต้องไม่สร้างซ้ำ"
        assert students == 1, "student ต้องไม่สร้างซ้ำ"
        # ข้อมูลถูกอัปเดตเป็นเวอร์ชันใหม่
        row = await conn.fetchrow(
            "SELECT s.last_name, s.class_role FROM students s WHERE s.student_id = '47001'"
        )
        assert row["last_name"] == "เปลี่ยนนามสกุล"
        assert row["class_role"] == "class_president"


# === Section 6: GET /import-jobs (list + progress) ===

@pytest.mark.asyncio
async def test_list_import_jobs_shows_progress(client, db_pool, admin_user):
    """GET /import-jobs → มีงาน + progress_percent=100 หลัง worker เสร็จ"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])
    await import_service.process_import_job(db_pool, job_id)

    res = client.get("/api/import-jobs", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 200
    jobs = res.json()
    assert len(jobs) == 1
    assert jobs[0]["id"] == job_id
    assert jobs[0]["status"] == "COMPLETED"
    assert jobs[0]["progress_percent"] == 100
    # file_path ต้องไม่รั่วออกไป (internal)
    assert "file_path" not in jobs[0]
    assert "error_logs" in jobs[0]  # ฟิลด์ที่ frontend ใช้ได้


@pytest.mark.asyncio
async def test_list_import_jobs_orders_newest_first(client, admin_user):
    """2 งาน → เรียงใหม่สุดก่อน"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    id1 = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])
    id2 = await upload_and_get_job(client, token, [["47002", "ม.4/1", 2, "นาย", "คนสอง", "สอง", "", ""]])

    res = client.get("/api/import-jobs", headers={"Authorization": f"Bearer {token}"})
    jobs = res.json()
    assert jobs[0]["id"] == id2
    assert jobs[1]["id"] == id1


@pytest.mark.asyncio
async def test_list_import_jobs_requires_auth(client):
    """ไม่มี token → 401"""
    res = client.get("/api/import-jobs")
    assert res.status_code == 401


@pytest.mark.asyncio
async def test_list_import_jobs_forbidden_for_student(client, student_user):
    """นักเรียนธรรมดา → 403"""
    _, username, password = student_user
    token = login_token(client, username, password)
    res = client.get("/api/import-jobs", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 403


# === Section 7: Recovery — งานค้าง (worker เริ่มต้น) ===

@pytest.mark.asyncio
async def test_recover_stuck_processing_and_queued_jobs(client, db_pool, admin_user):
    """งานค้าง PROCESSING/QUEUED เกิน 35 นาที → recover กลับ QUEUED + ยิงคิวใหม่ (Deep DB + audit)"""
    _, username, password = admin_user
    token = login_token(client, username, password)

    # สร้าง 2 งาน แล้วดันสถานะ + updated_at ให้แก่เกินกำหนด
    job1 = await upload_and_get_job(client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]])
    job2 = await upload_and_get_job(client, token, [["47002", "ม.4/1", 2, "นาย", "คนสอง", "สอง", "", ""]])
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            UPDATE student_import_jobs
            SET status = CASE WHEN id = $1 THEN 'PROCESSING' ELSE 'QUEUED' END,
                updated_at = NOW() - INTERVAL '60 minutes'
            WHERE id IN ($1, $2)
            """,
            job1, job2,
        )

    with patch("services.import_service.enqueue_import_job", new_callable=AsyncMock) as mock_enqueue:
        await import_service.recover_stuck_jobs(db_pool)

    assert mock_enqueue.await_count == 2

    # 🔍 Deep DB: ทั้งคู่กลับ QUEUED + มี audit RECOVER_IMPORT_JOB 2 รายการ
    async with db_pool.acquire() as conn:
        statuses = await conn.fetch(
            "SELECT status FROM student_import_jobs WHERE id IN ($1, $2) ORDER BY id",
            job1, job2,
        )
        assert [s["status"] for s in statuses] == ["QUEUED", "QUEUED"]
        recovered = await conn.fetchval(
            "SELECT count(*) FROM audit_logs WHERE action = 'RECOVER_IMPORT_JOB'"
        )
        assert recovered == 2


# === Section 8: Review fixes — privilege escalation / scope / limits / fallback ===

@pytest.mark.asyncio
async def test_level_teacher_cannot_import_school_wide_roles(client, db_pool):
    """🛡️ ครูระดับชั้น ใส่ตำแหน่ง 'แอดมิน' ใน Excel → แถวนั้นถูกข้าม (กัน privilege escalation)"""
    username = f"tch{random.randint(100000, 999999)}"
    uid = await auth_service.register_user(
        db_pool, username, "1234", "ครู เทส", username, "", 0, "teacher"
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE students SET staff_level = 'ม.4' WHERE user_id = $1", uid)

    token = login_token(client, username, "1234")
    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", "แอดมิน"],       # 🛡️ ต้องถูกข้าม
        ["47002", "ม.4/1", 2, "นางสาว", "สมหญิง", "รักเรียน", "", "หัวหน้าห้อง"],  # อนุญาต
    ]
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 1
    assert result["skipped"] == 1

    # 🔍 Deep DB: ต้องไม่มี user/student 47001 เลย (แอดมินที่ครูสร้างต้องไม่ถูกสร้าง)
    async with db_pool.acquire() as conn:
        admin_student = await conn.fetchval(
            "SELECT id FROM students WHERE student_id = '47001' AND deleted_at IS NULL"
        )
        admin_user_row = await conn.fetchval(
            "SELECT id FROM users WHERE username = '47001' AND deleted_at IS NULL"
        )
        assert admin_student is None, "ครูระดับชั้นต้องไม่สร้างแอดมิน"
        assert admin_user_row is None
        # 47002 (หัวหน้าห้อง) ถูกสร้างปกติ
        assert await conn.fetchval(
            "SELECT count(*) FROM students WHERE student_id = '47002'"
        ) == 1


@pytest.mark.asyncio
async def test_school_wide_admin_can_import_admin_role(client, db_pool, admin_user):
    """แอดมิน (school-wide) นำเข้าตำแหน่ง 'แอดมิน' ได้ปกติ"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [["47001", "", 0, "นาย", "สมชาย", "ใจดี", "", "แอดมิน"]]  # ไม่ผูกห้อง
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 1

    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT class_role, is_admin FROM students WHERE student_id = '47001'"
        )
        assert row["class_role"] == "admin"
        assert row["is_admin"] is True


@pytest.mark.asyncio
async def test_teacher_without_level_cannot_upload(client, db_pool):
    """ครูที่ยังไม่มีระดับชั้น (scope='none') → 403 (ไม่งั้นนำเข้าทั้งโรงเรียน)"""
    username = f"tch{random.randint(100000, 999999)}"
    await auth_service.register_user(db_pool, username, "1234", "ครู เทส", username, "", 0, "teacher")

    token = login_token(client, username, "1234")
    res = upload_excel(client, token, make_xlsx_bytes(rows=[["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]))
    assert res.status_code == 403, res.text
    assert "ระดับชั้น" in res.json()["detail"]


@pytest.mark.asyncio
async def test_upload_rejects_custom_default_password(client, db_pool, admin_user):
    """default_password กำหนดเอง (≠ 1234) → 400 (กันกำหนดรหัสที่แปลกปลอม)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    res = upload_excel(
        client, token,
        make_xlsx_bytes(rows=[["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]),
        default_password="hack123",
    )
    assert res.status_code == 400, res.text


@pytest.mark.asyncio
async def test_upload_rejects_header_only_file(client, db_pool, admin_user):
    """ไฟล์มีแต่หัวตาราง/แถวว่างล้วน → 400 (ไม่มีข้อมูลให้นำเข้า)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    res = upload_excel(client, token, make_xlsx_bytes(rows=[]))
    assert res.status_code == 400, res.text
    assert "หัวตาราง" in res.json()["detail"]


@pytest.mark.asyncio
async def test_upload_rejects_over_max_rows(client, db_pool, admin_user):
    """ไฟล์เกิน IMPORT_MAX_ROWS (5000) → 400 (กัน DoS ไฟล์ยักษ์)"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""] for _ in range(5001)]
    res = upload_excel(client, token, make_xlsx_bytes(rows=rows))
    assert res.status_code == 400, res.text
    assert "5000" in res.json()["detail"]


@pytest.mark.asyncio
async def test_worker_batch_failure_fallback_no_double_count(client, db_pool, admin_user):
    """
    Batch พังกลาง chunk (room_code ยาวเกิน VARCHAR(10)) → fallback ทีละแถว
    🛡️ กันตัวนับ imported/skipped นับซ้ำจาก batch ที่ rollback ไปแล้ว
    """
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [
        ["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""],
        ["47002", "ม.4/1นี้ยาวเกินมากเกินไป", 2, "นาย", "คนกลาง", "พัง", "", ""],  # room_code >10 chars → batch fail
        ["47003", "ม.4/1", 3, "นาย", "คนสุดท้าย", "โอเค", "", ""],
    ]
    job_id = await upload_and_get_job(client, token, rows)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"
    assert result["imported"] == 2, f"ต้องได้ 2 (ไม่นับซ้ำจาก batch ที่ rollback): {result}"
    assert result["skipped"] == 1

    # 🔍 Deep DB: มี student 47001 + 47003 เท่านั้น
    async with db_pool.acquire() as conn:
        ids = [r["student_id"] for r in await conn.fetch(
            "SELECT student_id FROM students WHERE student_id IN ('47001','47002','47003') ORDER BY student_id"
        )]
        assert ids == ["47001", "47003"]


@pytest.mark.asyncio
async def test_level_teacher_list_and_start_scoped(client, db_pool, admin_user):
    """ครูระดับชั้น เห็น/เริ่มงานได้เฉพาะของระดับชั้นตัวเอง (scope filtering)"""
    _, username, password = admin_user
    admin_token = login_token(client, username, password)

    # แอดมินสร้างงาน (allowed_level = NULL)
    admin_job_id = await upload_and_get_job(
        client, admin_token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]
    )

    # ครู ม.4 สร้างงาน (allowed_level = 'ม.4')
    t_username = f"tch{random.randint(100000, 999999)}"
    t_uid = await auth_service.register_user(
        db_pool, t_username, "1234", "ครู ม.4", t_username, "", 0, "teacher"
    )
    async with db_pool.acquire() as conn:
        await conn.execute("UPDATE students SET staff_level = 'ม.4' WHERE user_id = $1", t_uid)
    t_token = login_token(client, t_username, "1234")
    teacher_job_id = await upload_and_get_job(
        client, t_token, [["47002", "ม.4/1", 2, "นางสาว", "สมหญิง", "รักเรียน", "", ""]]
    )

    # ครู ม.4 list → เห็นเฉพาะงานของตัวเอง (ไม่เห็นของแอดมิน)
    res = client.get("/api/import-jobs", headers={"Authorization": f"Bearer {t_token}"})
    assert res.status_code == 200
    job_ids = [j["id"] for j in res.json()]
    assert teacher_job_id in job_ids
    assert admin_job_id not in job_ids, "ครูระดับชั้นต้องไม่เห็นงานของแอดมิน"

    # ครู ม.4 start งานของแอดมิน → 403
    res = client.post(
        f"/api/start-import-job/{admin_job_id}",
        headers={"Authorization": f"Bearer {t_token}"},
    )
    assert res.status_code == 403, res.text


@pytest.mark.asyncio
async def test_worker_completed_clears_error_message(client, db_pool, admin_user):
    """🛡️ หลัง COMPLETED error_message ต้องว่าง (ไม่ค้าง error เก่าจากรอบที่แล้ว)"""
    admin_uid, username, password = admin_user
    token = login_token(client, username, password)
    rows = [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]
    job_id = await upload_and_get_job(client, token, rows)

    # ตั้ง error_message ค้างไว้ก่อน (จำลองรอบ FAILED เก่า)
    async with db_pool.acquire() as conn:
        await conn.execute(
            "UPDATE student_import_jobs SET error_message = 'error เก่า', status = 'FAILED' WHERE id = $1",
            job_id,
        )

    # เริ่มใหม่ → ตั้ง QUEUED (error ต้องถูกล้าง) → worker ทำเสร็จ
    with patch("services.import_service.enqueue_import_job", new_callable=AsyncMock):
        await import_service.start_import_job(
            db_pool, job_id,
            actor_user_id=admin_uid, access_scope="all",
        )
    await import_service.process_import_job(db_pool, job_id)

    async with db_pool.acquire() as conn:
        job = await conn.fetchrow("SELECT status, error_message FROM student_import_jobs WHERE id = $1", job_id)
        assert job["status"] == "COMPLETED"
        assert job["error_message"] is None, "error_message ต้องถูกล้างเมื่อสำเร็จ"


@pytest.mark.asyncio
async def test_reimport_admin_school_wide_no_duplicate(client, db_pool, admin_user):
    """นำเข้าแอดมิน (room_id NULL) ซ้ำ → ไม่สร้าง user/student ซ้ำ"""
    _, username, password = admin_user
    token = login_token(client, username, password)
    rows = [["47001", "", 0, "นาย", "สมชาย", "ใจดี", "", "แอดมิน"]]
    job_id = await upload_and_get_job(client, token, rows)
    await import_service.process_import_job(db_pool, job_id)

    job_id2 = await upload_and_get_job(client, token, rows)
    await import_service.process_import_job(db_pool, job_id2)

    async with db_pool.acquire() as conn:
        users = await conn.fetchval("SELECT count(*) FROM users WHERE username = '47001'")
        students = await conn.fetchval("SELECT count(*) FROM students WHERE student_id = '47001'")
        assert users == 1, "user ต้องไม่สร้างซ้ำ"
        assert students == 1, "student ต้องไม่สร้างซ้ำ"


# === Section 9: Template download (ไฟล์ตัวอย่างสำหรับ user) ===

@pytest.mark.asyncio
async def test_download_template_returns_valid_xlsx(client, admin_user):
    """GET /import-student-template → 200 + ไฟล์ .xlsx ที่เปิดได้ + หัวคอลัมน์ตรง KNOWN_COLUMNS"""
    _, username, password = admin_user
    token = login_token(client, username, password)

    res = client.get("/api/import-student-template", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # body เป็นไฟล์ zip (.xlsx) — ตรวจ magic bytes PK
    assert res.content[:2] == b"PK"

    # เปิดด้วย openpyxl ได้ + หัวคอลัมน์ต้องตรง KNOWN_COLUMNS (แหล่งเดียวกับ validator — กัน drift)
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["ข้อมูล"]
    headers = [c.value for c in ws[1] if c.value is not None]
    assert headers == import_service.KNOWN_COLUMNS
    # ตัวอย่าง 2 แถว (00001, 00002) อยู่จริง — เพื่อให้ user เห็นรูปแบบข้อมูล
    assert ws["A2"].value == "00001"
    assert ws["B3"].value == "ม.4/2"


@pytest.mark.asyncio
async def test_download_template_requires_auth(client):
    """ไม่มี token → 401"""
    res = client.get("/api/import-student-template")
    assert res.status_code == 401


@pytest.mark.asyncio
async def test_download_template_forbidden_for_student(client, student_user):
    """นักเรียนธรรมดา (ไม่มี MANAGE_STUDENTS) → 403"""
    _, username, password = student_user
    token = login_token(client, username, password)
    res = client.get("/api/import-student-template", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 403


# === Section 10: Audit log งานเสร็จ/ล้ม (bug fix — audit ต้องถูกเขียนใน transaction) ===

@pytest.mark.asyncio
async def test_worker_writes_complete_audit_in_transaction(client, db_pool, admin_user):
    """
    🛡️ worker จบงาน COMPLETED ต้องเขียน audit COMPLETE_IMPORT_JOB ใน transaction เดียว
    (เดิม audit ไปอยู่ใน except ของการลบไฟล์ → ตอนจบปกติไม่ถูกเขียน)
    """
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(
        client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]
    )

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "COMPLETED"

    # 🔍 Deep DB: ต้องมี audit COMPLETE_IMPORT_JOB + ไฟล์ถูกลบจาก storage
    async with db_pool.acquire() as conn:
        audit = await conn.fetchval(
            "SELECT count(*) FROM audit_logs WHERE action = 'COMPLETE_IMPORT_JOB' AND entity_id = $1",
            str(job_id),
        )
        assert audit == 1, "ต้องมี audit COMPLETE_IMPORT_JOB อย่างน้อย 1 รายการ"

    async with db_pool.acquire() as conn:
        path = await conn.fetchval("SELECT file_path FROM student_import_jobs WHERE id = $1", job_id)
    assert not os.path.exists(path), "ไฟล์ storage ต้องถูกลบเมื่อจบงาน COMPLETED"


@pytest.mark.asyncio
async def test_worker_writes_failed_audit_in_transaction(client, db_pool, admin_user):
    """
    🛡️ worker จบงาน FAILED ต้องเขียน audit FAIL_IMPORT_JOB ใน transaction เดียว
    (เดิม audit ถูกข้ามทั้งสองกรณี — กัน regression ทั้ง COMPLETED + FAILED)
    """
    _, username, password = admin_user
    token = login_token(client, username, password)
    job_id = await upload_and_get_job(
        client, token, [["47001", "ม.4/1", 1, "นาย", "สมชาย", "ใจดี", "", ""]]
    )

    # บังคับให้ล้ม: ลบไฟล์จาก storage แล้วเรียก process (เหมือน worker เจอไฟล์หาย)
    async with db_pool.acquire() as conn:
        path = await conn.fetchval("SELECT file_path FROM student_import_jobs WHERE id = $1", job_id)
    os.remove(path)

    result = await import_service.process_import_job(db_pool, job_id)
    assert result["status"] == "FAILED"

    # 🔍 Deep DB: ต้องมี audit FAIL_IMPORT_JOB + status='error' (ใน transaction เดียวกับสถานะ)
    async with db_pool.acquire() as conn:
        audit = await conn.fetchrow(
            "SELECT * FROM audit_logs WHERE action = 'FAIL_IMPORT_JOB' AND entity_id = $1",
            str(job_id),
        )
        assert audit is not None, "เมื่องาน FAILED ต้องมี audit FAIL_IMPORT_JOB"
        assert audit["status"] == "error"
        assert audit["error_detail"] is not None
